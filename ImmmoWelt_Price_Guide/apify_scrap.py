from apify_client import ApifyClient
import openpyxl
import json

def get_dataset_items(url, limit):
    # Initialize the ApifyClient with your API token
    client = ApifyClient("apify_api_ODNYdzGzgsEyN9NGAQ3vyPyqTuj1jq34zPry")

    # Prepare the Actor input
    run_input = {
        "startUrls": [{ "url": url }],
        "limit": limit,
        "proxyConfig": { "useApifyProxy": True },
    }

    # Run the Actor and wait for it to finish
    run = client.actor("b4d9dfcf9MzBnyW7z").call(run_input=run_input)

    # Get the output of the Actor run
    dataset_items = []
    for item in client.dataset(run["defaultDatasetId"]).iterate_items():
        dataset_items.append(item)

    return dataset_items

def write_data_to_excel(data_objects, file_name):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the headers
    headers = list(data_objects[0].keys())
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    # Write the data rows
    for row_num, data in enumerate(data_objects, 2):
        for col_num, header in enumerate(headers, 1):
            value = json.dumps(data.get(header, ""))  # Convert value to string, use empty string if key not found
            sheet.cell(row=row_num, column=col_num).value = value

    # Save the workbook
    workbook.save(file_name)